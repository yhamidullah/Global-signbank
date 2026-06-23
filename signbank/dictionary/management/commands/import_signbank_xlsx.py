"""Import glosses from SignBank.xlsx into a new (or existing) Signbank dataset.

Columns expected in the first sheet:
    Name | ID | Mouth | HamNoSys

Usage:
    python manage.py import_signbank_xlsx /path/to/SignBank.xlsx \\
        --dataset-acronym DGS \\
        --dataset-name "Deutsche Gebärdensprache" \\
        [--language-code en] [--dry-run] [--limit N]
"""

import re
import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from signbank.dictionary.models import (
    Dataset, Language, SignLanguage, Gloss,
    LemmaIdgloss, LemmaIdglossTranslation,
    AnnotationIdglossTranslation,
)

ANNOTATION_MAX = 30   # DB max_length on AnnotationIdglossTranslation.text
LEMMA_MAX = 30        # DB max_length on LemmaIdglossTranslation.text


def _strip_ilex_suffix(name):
    """Remove iLex phonetic suffixes: NAME_1A'bew'lok → NAME_1A"""
    return re.split(r"'", str(name))[0].strip()


def _lemma_base(annotation_text):
    """Strip trailing variant letter: NAME_1A → NAME_1"""
    return re.sub(r'[A-Z]$', '', annotation_text).rstrip('_')


class Command(BaseCommand):
    help = 'Import SignBank.xlsx glosses into a Signbank dataset (creates it if needed).'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_file', help='Path to SignBank.xlsx')
        parser.add_argument('--dataset-acronym', default='DGS',
                            help='Acronym for the dataset (default: DGS)')
        parser.add_argument('--dataset-name', default='Deutsche Gebärdensprache',
                            help='Full name for the dataset')
        parser.add_argument('--language-code', default='en',
                            help='2-char language code for the dataset (default: en)')
        parser.add_argument('--sign-language', default=None,
                            help='Name of SignLanguage to attach to the dataset '
                                 '(default: first SignLanguage in DB)')
        parser.add_argument('--dry-run', action='store_true',
                            help='Print what would be imported without writing')
        parser.add_argument('--limit', type=int, default=0,
                            help='Stop after N rows (0 = all, useful for testing)')
        parser.add_argument('--keep-apostrophes', action='store_true',
                            help='Keep glosses whose name contains an apostrophe '
                                 "(iLex phonetic variants like NAME_1A'bew). By "
                                 'default these rows are skipped, since they usually '
                                 'have no video and carry hard parsing rules.')

    def handle(self, *args, **options):
        dry_run = options['dry_run']
        limit = options['limit']
        keep_apostrophes = options['keep_apostrophes']
        acronym = options['dataset_acronym']
        ds_name = options['dataset_name']
        lang_code = options['language_code']

        # ── Language ────────────────────────────────────────────────────────
        language = Language.objects.filter(language_code_2char=lang_code).first()
        if not language:
            self.stderr.write(f'Language with code "{lang_code}" not found in DB.')
            return
        self.stdout.write(f'Language: {language.name} ({lang_code})')

        # ── SignLanguage ─────────────────────────────────────────────────────
        if options['sign_language']:
            sign_language = SignLanguage.objects.get(name=options['sign_language'])
        else:
            sign_language = SignLanguage.objects.first()
        self.stdout.write(f'Sign language: {sign_language.name}')

        # ── Dataset ─────────────────────────────────────────────────────────
        if not dry_run:
            dataset, created = Dataset.objects.get_or_create(
                acronym=acronym,
                defaults={'name': ds_name, 'signlanguage': sign_language},
            )
            if created:
                dataset.translation_languages.add(language)
                dataset.default_language = language
                dataset.save()
                self.stdout.write(f'Created dataset: {ds_name} ({acronym})')
            else:
                self.stdout.write(f'Using existing dataset: {dataset.name} ({acronym})')
                if language not in dataset.translation_languages.all():
                    dataset.translation_languages.add(language)
        else:
            self.stdout.write(f'[dry-run] Would use/create dataset: {ds_name} ({acronym})')
            dataset = Dataset.objects.filter(acronym=acronym).first()

        admin_user = User.objects.filter(is_superuser=True).first()

        # ── Load xlsx ────────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(options['xlsx_file'], read_only=True, data_only=True)
        ws = wb.active
        self.stdout.write(f'Sheet: {ws.title}  rows≈{ws.max_row}')

        created_count = skipped = errors = skipped_apostrophe = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_name = row[0]
            raw_id   = row[1]
            mouth    = str(row[2]).strip() if row[2] else ''
            hamnosys = str(row[3]).strip() if row[3] else ''

            if not raw_name or not raw_id:
                continue

            # By default skip iLex phonetic-variant names (those containing an
            # apostrophe, straight ' or curly ’). They usually have no video and
            # hard parsing rules. Use --keep-apostrophes to import them anyway.
            if not keep_apostrophes and ("'" in str(raw_name) or "’" in str(raw_name)):
                skipped_apostrophe += 1
                continue

            try:
                ilex_id = str(int(raw_id))
            except (TypeError, ValueError):
                continue

            # Use full name (preserves uniqueness across iLex variants).
            # Strip suffix only for the lemma grouping, not the annotation text.
            annotation_text = str(raw_name)[:ANNOTATION_MAX]
            lemma_text = _lemma_base(_strip_ilex_suffix(str(raw_name)))[:LEMMA_MAX]

            if dry_run:
                self.stdout.write(
                    f'  [{ilex_id}] {annotation_text}'
                    + (f'  mouth={mouth}' if mouth else '')
                )
                created_count += 1
                if limit and created_count >= limit:
                    break
                continue

            # Idempotent: skip if this ilex ID already exists in this dataset
            if Gloss.objects.filter(alternative_id=ilex_id, lemma__dataset=dataset).exists():
                skipped += 1
                continue

            try:
                # Lemma: reuse if same text already exists in dataset
                lemma_trans_qs = LemmaIdglossTranslation.objects.filter(
                    text=lemma_text, language=language, lemma__dataset=dataset
                )
                if lemma_trans_qs.exists():
                    lemma = lemma_trans_qs.first().lemma
                else:
                    lemma = LemmaIdgloss.objects.create(dataset=dataset)
                    LemmaIdglossTranslation.objects.create(
                        lemma=lemma, language=language, text=lemma_text
                    )

                gloss = Gloss.objects.create(
                    lemma=lemma,
                    alternative_id=ilex_id,
                    mouthing=mouth,
                    hamnosys=hamnosys,
                )
                if admin_user:
                    gloss.creator.add(admin_user)

                AnnotationIdglossTranslation.objects.create(
                    gloss=gloss,
                    language=language,
                    text=annotation_text,
                )

                created_count += 1
                if created_count % 500 == 0:
                    self.stdout.write(
                        f'  ... {created_count} created, {skipped} skipped, {errors} errors'
                    )

            except Exception as exc:
                self.stderr.write(f'  ERROR [{ilex_id}] {annotation_text}: {exc}')
                errors += 1

            if limit and (created_count + skipped) >= limit:
                break

        action = 'Would create' if dry_run else 'Created'
        self.stdout.write(
            f'\nDone — {action}: {created_count} | '
            f'skipped (exists): {skipped} | '
            f'skipped (apostrophe): {skipped_apostrophe} | errors: {errors}'
        )
